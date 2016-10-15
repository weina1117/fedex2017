__author__ = 'weinaguo'
import openpyxl
import pandas as pd

headers = ['weight', 'rate', 'zone', 'from', 'to', 'commitment', 'norm_name']

name_mapping = {
    'EnvelopeUpTo8oz': '1Day Early AM(D)_Envelope',
    'EnvelopeUpTo8ozp': '1Day AM(D)_Envelope',
    'EnvelopeUpTo8ozs': '1Day PM(D)_Envelope',
    'EnvelopeUpTo8oz2': '2Day AM(D)_Envelope',
    'ExpressPackageFirstOvernight': '1Day Early AM(D)_Package',
    'ExpressPackagePriorityOvernight': '1Day AM(D)_Package',
    'ExpressPackageStandardOvernight': '1Day PM(D)_Package',
    'ExpressPackage2DayAM': '2Day AM(D)_Package',
    'ExpressPackage2Day': '2Day PM(D)_Package',
    'ExpressPackageSaver': '3Day(D)_Package',
    'GroundAndHomeDelivery': 'Ground Comm(D)',
    'ExpressMultiweight': '1Day Early AM(D)_CWT',
    'ExpressMultiweight2': '1Day AM(D)_CWT',
    'ExpressMultiweight3': '1Day PM(D)_CWT',
    'ExpressMultiweight4': '2Day AM(D)_CWT',
    'ExpressMultiweight5': '2Day(D)_CWT',
    'ExpressMultiweight6': '3Day(D)_CWT',
    'FirstOvernightFreight2': '1Day Frt Early AM(D)',
    '1DayFreight': '1Day Frt(D)',
    '2DayFreight': '2Day Frt(D)',
    '3DayFreight': '3Day(D)',
    'SameDayFreight': 'Same Day Frt(D)',
    'SameDayPerShipment': 'samedaypershipment'
}

def envelopeupto8oz(wb):
    norm_name = 'EnvelopeUpTo8oz'
    env = wb.get_sheet_by_name(norm_name)
    firstovernight = [env['A'+str(x)].value for x in range(2, env.max_row+1)]     # List comprehension
    zone = [env['B'+str(x)].value for x in range(2, env.max_row+1)]
    ret = pd.DataFrame()
    ret['rate'] = firstovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def envelopeupto8ozp(wb):
    norm_name = 'EnvelopeUpTo8ozp'
    env = wb.get_sheet_by_name(norm_name)
    priorityovernight = [env['A'+str(x)].value for x in range(2, env.max_row+1)]
    zone = [env['B'+str(x)].value for x in range(2, env.max_row+1)]
    ret = pd.DataFrame()
    ret['rate'] = priorityovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def envelopeupto8ozs(wb):
    norm_name = 'EnvelopeUpTo8ozs'
    env = wb.get_sheet_by_name(norm_name)
    standardovernight = [env['A'+str(x)].value for x in range(2, env.max_row+1)]
    zone = [env['B'+str(x)].value for x in range(2, env.max_row+1)]
    ret = pd.DataFrame()
    ret['rate'] = standardovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def envelopeupto8oz2(wb):
    norm_name = 'EnvelopeUpTo8oz2'
    env = wb.get_sheet_by_name(norm_name)
    twodayam = [env['A'+str(x)].value for x in range(2, env.max_row+1)]
    zone = [env['B'+str(x)].value for x in range(2, env.max_row+1)]
    ret = pd.DataFrame()
    ret['rate'] = twodayam
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def pfirstovernight(wb):
    norm_name = 'ExpressPackageFirstOvernight'
    pfo = wb.get_sheet_by_name(norm_name)
    weight = [pfo['A'+str(x)].value for x in range(2, pfo.max_row+1)]
    firstovernight = [pfo['B'+str(x)].value for x in range(2, pfo.max_row+1)]
    zone = [pfo['C'+str(x)].value for x in range(2, pfo.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = firstovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def ppriorityovernight(wb):
    norm_name = 'ExpressPackagePriorityOvernight'
    ppo = wb.get_sheet_by_name(norm_name)
    weight = [ppo['A'+str(x)].value for x in range(2, ppo.max_row+1)]
    priorityovernight = [ppo['B'+str(x)].value for x in range(2, ppo.max_row+1)]
    zone = [ppo['C'+str(x)].value for x in range(2, ppo.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = priorityovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def pstandardovernight(wb):
    norm_name = 'ExpressPackageStandardOvernight'
    pso = wb.get_sheet_by_name(norm_name)
    weight = [pso['A'+str(x)].value for x in range(2, pso.max_row+1)]
    standardovernight = [pso['B'+str(x)].value for x in range(2, pso.max_row+1)]
    zone = [pso['C'+str(x)].value for x in range(2, pso.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = standardovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def ptwodayam(wb):
    norm_name = 'ExpressPackage2DayAM'
    p2a = wb.get_sheet_by_name(norm_name)
    weight = [p2a['A'+str(x)].value for x in range(2, p2a.max_row+1)]
    rate = [p2a['B'+str(x)].value for x in range(2, p2a.max_row+1)]
    zone = [p2a['C'+str(x)].value for x in range(2, p2a.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def ptwoday(wb):
    norm_name = 'ExpressPackage2Day'
    p2 = wb.get_sheet_by_name(norm_name)
    weight = [p2['A'+str(x)].value for x in range(2, p2.max_row+1)]
    rate = [p2['B'+str(x)].value for x in range(2, p2.max_row+1)]
    zone = [p2['C'+str(x)].value for x in range(2, p2.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def psaver(wb):
    norm_name = 'ExpressPackageSaver'
    ps = wb.get_sheet_by_name(norm_name)
    weight = [ps['A'+str(x)].value for x in range(2, ps.max_row+1)]
    rate = [ps['B'+str(x)].value for x in range(2, ps.max_row+1)]
    zone = [ps['C'+str(x)].value for x in range(2, ps.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def groundandhomedelivery(wb):
    norm_name = 'GroundAndHomeDelivery'
    ghd = wb.get_sheet_by_name(norm_name)
    weight = [ghd['A'+str(x)].value for x in range(2, ghd.max_row+1)]
    rate = [ghd['B'+str(x)].value for x in range(2, ghd.max_row+1)]
    zone = [ghd['C'+str(x)].value for x in range(2, ghd.max_row+1)]
    commitment = [ghd['D'+str(x)].value for x in range(2, ghd.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = rate
    ret['zone'] = zone
    ret['commitment'] = commitment
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def expressmultiweight(wb):
    norm_name = 'ExpressMultiweight'
    emw = wb.get_sheet_by_name(norm_name)
    ws = [emw['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in emw['A'+str(x)].value else [emw['A'+str(x)].value[:-1], 0]
          for x in range(3, emw.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [emw['B'+str(x)].value for x in range(3, emw.max_row+1)]
    zone = [emw['C'+str(x)].value for x in range(3, emw.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def expressmultiweight2(wb):
    norm_name = 'ExpressMultiweight2'
    emw = wb.get_sheet_by_name(norm_name)
    ws = [emw['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in emw['A'+str(x)].value else [emw['A'+str(x)].value[:-1], 0]
          for x in range(3, emw.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [emw['B'+str(x)].value for x in range(3, emw.max_row+1)]
    zone = [emw['C'+str(x)].value for x in range(3, emw.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def expressmultiweight3(wb):
    norm_name = 'ExpressMultiweight3'
    emw = wb.get_sheet_by_name(norm_name)
    ws = [emw['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in emw['A'+str(x)].value else [emw['A'+str(x)].value[:-1], 0]
          for x in range(3, emw.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [emw['B'+str(x)].value for x in range(3, emw.max_row+1)]
    zone = [emw['C'+str(x)].value for x in range(3, emw.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def expressmultiweight4(wb):
    norm_name = 'ExpressMultiweight4'
    emw = wb.get_sheet_by_name(norm_name)
    ws = [emw['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in emw['A'+str(x)].value else [emw['A'+str(x)].value[:-1], 0]
          for x in range(3, emw.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [emw['B'+str(x)].value for x in range(3, emw.max_row+1)]
    zone = [emw['C'+str(x)].value for x in range(3, emw.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def expressmultiweight5(wb):
    norm_name = 'ExpressMultiweight5'
    emw = wb.get_sheet_by_name(norm_name)
    ws = [emw['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in emw['A'+str(x)].value else [emw['A'+str(x)].value[:-1], 0]
          for x in range(3, emw.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [emw['B'+str(x)].value for x in range(3, emw.max_row+1)]
    zone = [emw['C'+str(x)].value for x in range(3, emw.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def expressmultiweight6(wb):
    norm_name = 'ExpressMultiweight6'
    emw = wb.get_sheet_by_name(norm_name)

    ws = [emw['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in emw['A'+str(x)].value else[emw['A'+str(x)].value[:-1], 0]
          for x in range(3, emw.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [emw['B'+str(x)].value for x in range(3, emw.max_row+1)]
    zone = [emw['C'+str(x)].value for x in range(3, emw.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def firstovernightfreight2(wb):
    norm_name = 'FirstOvernightFreight2'
    fof = wb.get_sheet_by_name(norm_name)
    ws = [fof['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in fof['A'+str(x)].value else [fof['A'+str(x)].value[:-1], 0]
          for x in range(2, fof.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [fof['B'+str(x)].value for x in range(2, fof.max_row+1)]
    zone = [fof['C'+str(x)].value for x in range(2, fof.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def onedayfreight(wb):
    norm_name = '1DayFreight'
    odf = wb.get_sheet_by_name(norm_name)
    ws = [odf['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in odf['A'+str(x)].value else [odf['A'+str(x)].value[:-1], 0]
          for x in range(2, odf.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [odf['B'+str(x)].value for x in range(2, odf.max_row+1)]
    zone = [odf['C'+str(x)].value for x in range(2, odf.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def twodayfreight(wb):
    norm_name = '2DayFreight'
    tdf = wb.get_sheet_by_name(norm_name)
    ws = [tdf['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in tdf['A'+str(x)].value else [tdf['A'+str(x)].value[:-1], 0]
          for x in range(2, tdf.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [tdf['B'+str(x)].value for x in range(2, tdf.max_row+1)]
    zone = [tdf['C'+str(x)].value for x in range(2, tdf.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def threedayfreight(wb):
    norm_name = '3DayFreight'
    thf = wb.get_sheet_by_name(norm_name)
    ws = [thf['A'+str(x)].value.split(u'\u2013')
          if u'\u2013' in thf['A'+str(x)].value else [thf['A'+str(x)].value[:-1], 0]
          for x in range(2, thf.max_row+1)]
    w_from, w_to = [x[0] for x in ws], [x[1] for x in ws]
    rate = [thf['B'+str(x)].value for x in range(2, thf.max_row+1)]
    zone = [thf['C'+str(x)].value for x in range(2, thf.max_row+1)]

    ret = pd.DataFrame()
    ret['from'] = w_from
    ret['to'] = w_to
    ret['rate'] = rate
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def samedayfreight(wb):
    norm_name = 'SameDayFreight'
    pso = wb.get_sheet_by_name(norm_name)
    weight = [pso['A'+str(x)].value for x in range(2, pso.max_row+1)]
    standardovernight = [pso['B'+str(x)].value for x in range(2, pso.max_row+1)]
    zone = [pso['C'+str(x)].value for x in range(2, pso.max_row+1)]
    ret = pd.DataFrame()
    ret['weight'] = weight
    ret['rate'] = standardovernight
    ret['zone'] = zone
    ret['norm_name'] = name_mapping[norm_name]
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

def sameday():
    ret = pd.DataFrame()
    weight = [x for x in range(70)]
    rate = [235 if x < 25 else 235+(x-24)*1.35 for x in weight]
    ret['weight'] = weight
    ret['rate'] = rate
    ret['norm_name'] = name_mapping['SameDayPerShipment']
    for h in headers:
        if h not in ret:
            ret[h] = [0 for _ in range(len(ret['rate']))]
    return ret

rules = {
    'envelopeupto8oz': envelopeupto8oz,
    'envelopeUpTo8ozp': envelopeupto8ozp,
    'envelopeUpTo8ozs': envelopeupto8ozs,
    'envelopeUpTo8oz2': envelopeupto8oz2,
    'pfirstovernight': pfirstovernight,
    'ppriorityovernight': ppriorityovernight,
    'pstandardovernight': pstandardovernight,
    'ptwodayam': ptwodayam,
    'ptwoday': ptwoday,
    'psaver': psaver,
    'groundandhomedelivery': groundandhomedelivery,
    'expressmultiweight': expressmultiweight,
    'expressmultiweight2': expressmultiweight2,
    'expressmultiweight3': expressmultiweight3,
    'expressmultiweight4': expressmultiweight4,
    'expressmultiweight5': expressmultiweight5,
    'expressmultiweight6': expressmultiweight6,
    'firstovernightfreight2': firstovernightfreight2,
    'onedayfreight': onedayfreight,
    'twodayfreight': twodayfreight,
    'threedayfreight': threedayfreight,
    'samedayfreight': samedayfreight,
    'sameday': samedayfreight

}



def main():
    wb = openpyxl.load_workbook('/Users/weinaguo/Desktop/2017fedexrates.xlsx')
    results = []
    for rule in rules:
        results.append(rules[rule](wb))

    result = pd.DataFrame()
    for r in results:
        result = result.append(r)
    writer = pd.ExcelWriter('/Users/weinaguo/Desktop/fedex2017-out.xlsx')
    result.to_excel(writer)
    writer.save()

if __name__ == '__main__':
    main()
